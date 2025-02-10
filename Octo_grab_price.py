# 	TO DO List
# 1. Find a way to stop loading webpages once the all the tables are loaded. Currently
	   # each loop takes about 11-15 seconds. If that can be cut in half, it would run
	   # twice as fast... I made a funny.
# 2.  
# 	Grab Prices from Octopart & add it to existing Spreadsheet
# DESCRIPTION:
# This script uses the MPN and Kitted Qty information in a spreadsheet to
# search Octopart, find the lowest cost per part, and add the cost to the 
# Cost column in the spreadsheet. The script also 

# WEB ADDRESS: https://octopart.com/search?q=CRCW12064M70FKEA&currency=USD&specs=0
# MPN: CRCW12064M70FKEA, Kitted Qty: 6 (<10), Cost each: $0.027 (s/b $0.03)
# MPN: LT8645SEV#PBF	https://octopart.com/search?q=LT8645SEV%23PBF&currency=USD&specs=0

# ALGORITHM:
# 1. Open Excel, get the locations of usable columns.
# 2. Create a link to Octopart.com for the MPN, create a dataframe for the  
	# Distributor/Qty table.
	# 2.1 If the Press & Hold screen comes up, take care of it & move on.
	# 2.2 If the table doesn't exist, try to use the generic price if possible.
# 3. Match Kitted Qty with right Octopart Qty column (i.e. 100 > 49 >10).
# 4. Extract lowest price and add it to Excel. Round up to nearest penny.
# 5. Save Excel, continue for each MPN until finished.
# 6. Report errors, so the user knows all the data was found.

# Instructions: Place the spreadsheet you want to work on inside this folder.
			  # Change the 'fname' below to the filename of the spreadsheet
			  # you want to modify.'''

fname = 'Updated_BOM_TDU3.0_04AUG2021_byQTY.xlsx'

# Optional: Open the spreadsheet and empty the cells in the Cost Each column. 
# Close the spreadsheet listed above. The script will open and close Chrome.
# The computer will be partially unusable during this time because the script
# will open/close Chrome which occupies about 1/2 of the monitor.'''


print('\nStarting OctoPart Price Grab Script\n')

import glob
import math
import os
import re
import time
import zipfile

import openpyxl
import pandas as pd
import pyautogui
from   selenium import webdriver
from   selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from   selenium.webdriver.support.ui import WebDriverWait
from   selenium.webdriver.support    import expected_conditions as EC
from   selenium.webdriver.common.by  import By
# Necessary modules not listed are: 	lxml, opencv-python, pillow

import honlogger


# Honeywell Logger constants
DEPT         = 'TITAN'
SCRIPT_NAME  = (os.path.split(__file__)[1]).split('.')[0]	# Find the script's filename.
FILENAMES    = [SCRIPT_NAME, fname]
LOG_START    = [time.time(), time.ctime()]
NOUN_VERB    = "Labels Made"
MIN_PER_ITEM = 1.0

# Honeywell Logger variable
total_items  = 0


# Provide a human readable view of the data.
def ordered_preview(lol):
	LineOfLabels = []
	count = 0
	for label in lol:
		LineOfLabels.append(label)
		count += 1
		if count == 2:
			print(LineOfLabels)
			LineOfLabels = []
			count = 0


def defeat_press_n_hold():
	seenPNH = 0
	x, y    = 0, 0
	time.sleep(3)
	print('Trying to defeat the press & hold system!!!')
	while True:
		print("Current seenPNH value:", seenPNH)
		try:  # Looking for hollow oval.
			pyautogui.screenshot('aaa_PNH-1hollow-oval.png')
			x, y = pyautogui.locateCenterOnScreen('.\Drivers\Press-N-Hold.jpg', confidence=0.7)  # Takes ~1s to locate.
		except:  # Button not found.
			if seenPNH == 0:
				pyautogui.screenshot('aaa_PNH-11hollow-oval.png')
				print("              BUTTON NOT FOUND!!! If you see it, something is wrong...")
				return 0
			elif seenPNH == 1:
				print("Button was found. Press & Hold defeated!")
				return 1
		print('BUTTON FOUND!!! Pressing your buttons!')
		pyautogui.moveTo(x, y)
		pyautogui.drag(25, 0, 5, button='left')  # Press button
		seenPNH = 1
		pyautogui.screenshot('aaa_PNH-2hollow-pressed.png')
		time.sleep(1)

		# Looking for black oval.
		while True:
			pyautogui.screenshot('aaa_PNH-3black-oval.png')
			try:
				x, y = pyautogui.locateCenterOnScreen('.\Drivers\Press-N-Hold2_rt.jpg', confidence=0.7)
				print('Seeing button full.')
			except:
				pyautogui.screenshot('aaa_PNH-4error image.png')
				print("Didn't see button full.")
				break
			pyautogui.moveTo(x, y)
		time.sleep(1)


def make_list_from_sheet(fname):
	# OPEN THE WORKBOOK and sheet that contains the Master Item Number.
	wb = openpyxl.load_workbook(fname)
	sheet = wb.worksheets[0]
	TopLeft = 'A1'
	maxCol = openpyxl.utils.get_column_letter(sheet.max_column)
	BotRght = maxCol + str(sheet.max_row)
	
	# 1. Open Excel, get the locations of the right columns. and get the data
	Columnsdict = {'kitted qty': 'empty', 'manufacturer part number': 'empty',
				   'honeywell part description': 'empty', 'cost each': 'empty'}
	#Columnsdict = {'kitted qty': 'empty', 'manufacturer part number': 'empty',  Was the original
	#			   'honeywell part description': 'empty', 'cost each': 'empty'}	 dictionary
	print("Finding Relevant Columns.")
	for rowOfCellObjects in sheet['A1':maxCol + '10']: # Looking at the title row in excel.
		if 'empty' not in Columnsdict.values(): # All columns have been found.
			print("TOP LEVEL. Finished searching. BREAK!")
			print("Topleft is now", TopLeft)
			break
		for cellObj in rowOfCellObjects:
			try:	# Fixes issue when 'None' is in the cell value.
				search_key = cellObj.value.lower()
				print(search_key)
			except:
				continue
			single_dict = dict(filter(lambda item: search_key in item[0], Columnsdict.items()))
			for key in single_dict:
				Columnsdict[key] = cellObj.coordinate[:-1]
				TopLeft = cellObj.coordinate
				TopLeft = int(re.split('(\d+)', TopLeft)[1]) + 1
				TopLeft = 'A' + str(TopLeft)
	print("Relevant columns are:", Columnsdict)
	print("Generating Data tables from spreadsheet.")
	SheetContents = []  # ['65', 'SN74LVC244ARGYR', 'MICROCIRCUIT, LOGIC,']
	MPN_Cost_list = []  # ['SN74LVC244ARGYR', 'Y7']
	for rowOfCellObjects in sheet[TopLeft:BotRght]:
		MPN_Cost_item = []
		sublist = []
		kQty = ''
		MPN  = ''
		HPD  = ''
		cost = ''
		for cellObj in rowOfCellObjects: # Take in the data even if it's out of order.
			colLetter = cellObj.column_letter
			if colLetter in Columnsdict.values():	# Sort everything into the correct order.
				if colLetter == Columnsdict['kitted qty']:
					kQty = cellObj.value
				elif colLetter == Columnsdict['manufacturer part number']:
					MPN = cellObj.value
					MPN_Cost_item.append(cellObj.value)
					MPN_Cost_item.append(Columnsdict['cost each'] + str(cellObj.row))
				elif colLetter == Columnsdict['honeywell part description']:
					HPD = cellObj.value
				elif colLetter == Columnsdict['cost each']:
					cost = cellObj.value
		MPN_Cost_list.append(MPN_Cost_item)
		sublist.append(kQty)
		sublist.append(MPN)
		sublist.append(HPD)
		sublist.append(cost)
		SheetContents.append(sublist)
	wb.close()
	return SheetContents, MPN_Cost_list


def remove_nones(ListContents):
	print("Cleaning up the data to make it easier to work with.")
	ListContents2 = []
	for sublist in ListContents:  # Get rid of Nones
		cleaned = [elem for elem in sublist if elem is not None]
		if len(cleaned):
			ListContents2.append(cleaned)
	ListContents = ListContents2
	return ListContents
	

def convert_to_strings(ListContents, elem_min):
	# Convert all items in list to strings & that each list is a 
	# minimum size. Report bad rows that need corrected. Do not halt.
	ListContents2 = []
	for index, sublist in enumerate(ListContents):
		if len(sublist) >= elem_min:  # Should have > 3  items in each list/row.
			pass
		else:
			print("Something is wrong with the spreadsheet on line: ", index + 3)  # Data starts on Row 3.
			continue
		cleaned = [str(elem) for elem in sublist] # Make everything a string.
		ListContents2.append(cleaned)
	ListContents = ListContents2
	return ListContents


def formula_to_int(ListContents):
	# Convert any Kitted Qty formulas into a total Qty.
	# Example: '=537+913+300' will become '1750'
	ListContents2 = []
	for sublist in ListContents:
		list = []
		kitted_qty = (sublist[0]).replace('=', '')
		if kitted_qty.lower() == 'n/a':
			kitted_qty = '0'
		if kitted_qty.lower() == '':
			continue
		kitted_qty = sum(int(i) for i in kitted_qty.split('+'))
		list.append(str(kitted_qty))
		list.append(sublist[1])
		list.append(sublist[2])
		#list.append([cost, sublist[1], sublist[2]])	# Previous code.
		ListContents2.append(list)
	ListContents = ListContents2
	return ListContents


def cleanup_list(ListContents):
	# Clean up the list, so it's easier to view.
	print("Getting data ready for preview.\n")
	ListContents2 = []
	for sublist in ListContents:  # Fix MPN & description items.
		if len(sublist) == 1:
			break
		item_list = []
		for index, item in enumerate(sublist):
			if index == 0 or index == 1:  # Grab first MPN, eliminate rest from list
				try:
					item = item.split('\n')[0]
					item = item.split(',')[0]
				except:
					print(sublist)
					print("I'm quitting because something went wrong!!!")
					quit()
			elif index == 2:  # Shorten Description to 20 characters.
				item = item[:20]
			item_list.append(item)
		ListContents2.append(item_list)
	ListContents = ListContents2
	print(ordered_preview(ListContents))
	return ListContents


def compress_artifacts():
	# Take all the .png files, the excel file and zip it up
	art_list = glob.glob('*.png') 
	art_list.append(fname)
	zip_file = fname.split('.')[0]+'.zip'
	print("\nCopying screenshots and excel file into a zip file listed below")
	print(zip_file)
	with zipfile.ZipFile((zip_file), 'w') as zipMe:
		for file in art_list:
			zipMe.write(file, compress_type=zipfile.ZIP_DEFLATED)
	art_list = art_list[:-1]
	print("\nDeleting all screenshots that were copied to the zip.\n")
	for file in art_list:
		os.remove(file)
		

################### MAIN PROGRAM  SECTION ##############################
SheetContents, MPN_Cost_list = make_list_from_sheet(fname)

MPN_Cost_list = remove_nones(MPN_Cost_list)
SheetContents = remove_nones(SheetContents)

MPN_Cost_list = convert_to_strings(MPN_Cost_list, 2)
SheetContents = convert_to_strings(SheetContents, 3)

SheetContents = formula_to_int(SheetContents)

MPN_Cost_list = cleanup_list(MPN_Cost_list)
SheetContents = cleanup_list(SheetContents)

total_MPNs = len(SheetContents)


print("\nThe order for the list above is the following:")
print("    ['Kitted Qty', 'Manufacturer Part Number', 'Brief Description']\n")

print('Done with spreadsheet. The number of MPNs is', total_MPNs)
answer = (input("Does the above look okay (Y/N)? ")).upper()

while True:  # Get the user to verify the SheetContents look okay.
	if answer == "Y":
		break
	elif answer == "N":
		quit()
	else:
		answer = (input("Try again. Choose Y or N: ")).upper()
print("Thank you, We will start getting data from Octopart.com now.")
time.sleep(5)

###########    STARTING MAIN PROGRAM LOOP     ##########################
count = 0
error = 0
wb    = openpyxl.load_workbook(fname)
sheet = wb[wb.sheetnames[0]]

for MPN, coord in MPN_Cost_list:
	count  += 1
	cost_ea = 0
	kitted_qty = 0
	print("\nThis is MPN #", count, "/", len(SheetContents))
	print("The MPN is:", MPN)

	for item in SheetContents:  # Get the Kitted Qty from SheetContents list.
		SheetMPN = item[1]
		if SheetMPN == MPN:
			kitted_qty = item[0]
			break

	if int(kitted_qty) < 1:
		print('The kitted Qty is zero, therefore the cost is ZERO!!!')
		sheet[coord] = 0
		wb.save(fname)
		continue

	options = webdriver.ChromeOptions()
	options.add_experimental_option('excludeSwitches', ['enable-logging'])
	driver = webdriver.Chrome(".\Drivers\chromedriver.exe", options=options)
	
	MPNurl = ((MPN.replace('#', '%23')).replace('/', '%2F')).replace(':', '%3A')
	MPNurl = "https://octopart.com/search?q=" + MPNurl + "&currency=USD&specs=0"
	print("The webpage is:", MPNurl)
	
	driver.get(MPNurl)
	driver.set_window_size(1200, 1000)
	
	webpage = driver.page_source
	start_table = webpage.find('<table class')
	end_table = webpage.find('</table>') + 8
	table = webpage[start_table:end_table]
	rough_octoprice = webpage[(webpage.find('currency">USD<')):]
	rough_octoprice = rough_octoprice[(rough_octoprice.find('price">') + 7):]

	if len(rough_octoprice) > 0:
		rough_octoprice = rough_octoprice.replace(',', '')
		rough_octoprice = float(rough_octoprice.split('<')[0])

	try:
		df = pd.read_html(table)
	except:
		print("\n\n\n\nSomething went wrong and I'm trying to fix it.\n\n")
		find_button = defeat_press_n_hold()  # try the press & hold strategy.
		if find_button == 1:  # Press & Hold button found & pressed, we believe page is good.
			table = driver.page_source
			start_table = table.find('<table class')
			end_table = table.find('</table>') + 8
			table = table[start_table:end_table]
			df = pd.read_html(table)
		elif find_button == 0:  # Press hold, not found, we believe no table exists.
			try:
				cost_ea = math.ceil(rough_octoprice.min() * 100) / 100
			except:
				print("\n        SOMETHING WENT WRONG AND NO TABLE AND NO ROUGH PRICE WER FOUND!!!")
				print("        SETTING COST FOR THIS ONE AT 'ERROR', SO YOU CAN FIX IT LATER.")
				cost_ea = 'ERROR'
				error = 1
			print('Table was not found for this MPN, but Generic cost was found on page.')
			print('The cost each is:', cost_ea)
	sname = (((MPN.replace('/', '')).replace(':', '')).replace("'", "")).replace('#', '')
	driver.save_screenshot(sname + '.png')
	print('The kitted Qty is:', kitted_qty)
	if cost_ea == 0:  # We didn't have an event where the generic price is the only option.
		driver.execute_script("window.stop();")
		driver.close()
		desired_col = ['1', '10', '100', '1,000', '10,000']
		desired_col2 = []
		for item in desired_col:  # Get the correct Qty column
			if int(kitted_qty) >= int(item.replace(',', '')):
				desired_col2.append(item)
		desired_col = desired_col2[len(desired_col2) - 1]
		df = df[0]
		try:
			cost_ea = math.ceil(df[desired_col].min() * 100) / 100
		except ValueError:
			print('Error occurred. Could not find Price. NaN found')
			cost_ea = 'ERROR'
			error = 1
		print('The desired column is:', desired_col)
	print('The cost each is:', cost_ea)
	try:
		print(df)
	except:
		pass
	print("\n")
	sheet[coord] = cost_ea
	wb.save(fname)

if error == 1:
	print("There was an ERROR when the script was running. Please check", fname)

compress_artifacts()

# Run Honeywell logger to write logs to file for YOY cost savings.
total_items = total_MPNs
honlogger.write_log(DEPT, FILENAMES, LOG_START, NOUN_VERB, total_items, MIN_PER_ITEM)	
