#############   OCTOPRICE TO SAP SCRIPT   ###############
# 1. Need to open spreadsheet
# 2. Find all the MPN, Desc, Qty, Cost and put it in a list of lists.
# 3. Clean up the list like normal. Anything that has a Qty  or Cost of 0, will be removed.
# 4. Switch over to SAP (alt-tab), then start pasting in the data along with 'US' for the country code.
# 5. Tell the user we're finished.


import glob
import math
import os
import re
import time
import zipfile

import openpyxl
import pyautogui

import honlogger


# Honeywell Logger variable
total_items  = 1
fname = ''

# Honeywell Logger constants
DEPT         = 'TITAN'
SCRIPT_NAME  = (os.path.split(__file__)[1]).split('.')[0]	# Find the script's filename.
FILENAMES    = [SCRIPT_NAME, fname]
LOG_START    = [time.time(), time.ctime()]
NOUN_VERB    = "Rows Entered"
MIN_PER_ITEM = 0.5


def find_fname():
	fhand = open('Octo_grab_price.py')
	for line in fhand:
		if line.startswith('fname'):
			print(line[9:-2])
			fname = line[9:-2]
			break
	return fname

def make_list_from_sheet(fname):
	# OPEN THE WORKBOOK and sheet that contains the Master Item Number.
	wb = openpyxl.load_workbook(fname)
	sheet = wb.worksheets[0]
	TopLeft = 'A1'
	maxCol = openpyxl.utils.get_column_letter(sheet.max_column)
	BotRght = maxCol + str(sheet.max_row)
	
	# 1. Open Excel, get the locations of the right columns. and get the data
	Columnsdict = {'kitted qty': 'empty', 'manufacturer part number': 'empty',
				   'honeywell part description': 'empty', 'cost each': 'empty'}
	print("Finding Relevant Columns.")
	for rowOfCellObjects in sheet['A1':maxCol + '10']: # Looking at the title row in excel.
		if 'empty' not in Columnsdict.values(): # All columns have been found.
			break
		for cellObj in rowOfCellObjects:
			try:	# Fixes issue when 'None' is in the cell value.
				search_key = cellObj.value.lower()
			except:
				continue
			single_dict = dict(filter(lambda item: search_key in item[0], Columnsdict.items()))
			for key in single_dict:
				Columnsdict[key] = cellObj.coordinate[:-1]
				TopLeft = cellObj.coordinate
				TopLeft = int(re.split('(\d+)', TopLeft)[1]) + 1
				TopLeft = 'A' + str(TopLeft)
	print("Relevant columns are:", Columnsdict)
	print("\nGenerating Data tables from spreadsheet.\n")
	SheetContents = []  # ['65', 'SN74LVC244ARGYR', 'MICROCIRCUIT, LOGIC,']
	for rowOfCellObjects in sheet[TopLeft:BotRght]:
		sublist = []
		kQty = ''
		MPN  = ''
		HPD  = ''
		cost = ''
		for cellObj in rowOfCellObjects: # Take in the data even if it's out of order.
			colLetter = cellObj.column_letter
			if colLetter in Columnsdict.values():	# Sort everything into the correct order.
				if colLetter == Columnsdict['kitted qty']:
					kQty = cellObj.value
				elif colLetter == Columnsdict['manufacturer part number']:
					MPN = cellObj.value
				elif colLetter == Columnsdict['honeywell part description']:
					HPD = cellObj.value
				elif colLetter == Columnsdict['cost each']:
					cost = cellObj.value
		sublist.append(kQty)
		sublist.append(MPN)
		sublist.append(HPD)
		sublist.append(cost)
		SheetContents.append(sublist)
	wb.close()
	return SheetContents


def remove_nones(ListContents):
	print("Cleaning up the data to make it easier to work with.")
	ListContents2 = []
	for sublist in ListContents:  # Get rid of Nones
		cleaned = [elem for elem in sublist if elem is not None]
		if len(cleaned):
			ListContents2.append(cleaned)
	ListContents = ListContents2
	return ListContents
	

def convert_to_strings(ListContents, elem_min):
	# Convert all items in list to strings & that each list is a 
	# minimum size. Report bad rows that need corrected. Do not halt.
	ListContents2 = []
	for index, sublist in enumerate(ListContents):
		if len(sublist) >= elem_min:  # Should have > 3  items in each list/row.
			pass
		else:
			print("Something is wrong with the spreadsheet on line: ", index + 3)  # Data starts on Row 3.
			continue
		cleaned = [str(elem) for elem in sublist] # Make everything a string.
		ListContents2.append(cleaned)
	ListContents = ListContents2
	return ListContents


def formula_to_int(ListContents):
	# Convert any Kitted Qty formulas into a total Qty.
	# Example: '=537+913+300' will become '1750'
	ListContents2 = []
	for sublist in ListContents:
		list = []
		kitted_qty = (sublist[0]).replace('=', '')
		if kitted_qty.lower() == 'n/a':
			kitted_qty = '0'
		if kitted_qty.lower() == '':
			continue
		kitted_qty = sum(int(i) for i in kitted_qty.split('+'))
		# If there's issue where it says 'Index out of range', then data may be missing.
		list.append(str(kitted_qty))
		list.append(sublist[1])
		list.append(sublist[2])
		list.append(sublist[3])
		ListContents2.append(list)
	ListContents = ListContents2
	return ListContents


def cleanup_list(ListContents):
	# Clean up the list, so it's easier to view.
	ListContents2 = []
	for sublist in ListContents:  # Fix MPN & description items.
		zero = 0
		item_list = []
		for index, item in enumerate(sublist):
			if index == 0: 	  # Check for and remove zero Qty items.
				if item == '0':
					zero = 1
					break
			if index == 1:    # Grab first MPN, eliminate rest from list
				try:
					item = item.split('\n')[0]
					item = item.split(',')[0]
				except:
					print(sublist)
					print("I'm quitting because something went wrong!!!")
					quit()
			elif index == 2:  # Shorten Description to 20 characters.
				item = item[:20]
			elif index == 3:  # Check for and remove zero cost items.
				if item == '0':
					zero = 1
					break
			item_list.append(item)
		if zero == 0:
			ListContents2.append(item_list)
	ListContents = ListContents2
	return ListContents


def copy_data_to_sap(ListContents):
	# Take the data & paste it into SAP. Ensure SAP has been selected
	# right before going back to the Editor.
	pyautogui.FAILSAFE = True	# Move mouse to corners to kill script!
	
	ALERT_TEXT = 'Move the mouse into any corner to kill the script.\n        Click OK to continue.'
	pyautogui.alert(text=ALERT_TEXT, title='!!!ALERT!!!', button='OK')
	time.sleep(3)
	
	# Switch from the command promt, to the SAP using ALT-TAB.
	pyautogui.keyDown('alt')
	pyautogui.typewrite(['tab','tab'], interval=1)	# If IDE is not open, 
	pyautogui.keyUp('alt')							# this should be edited.
	
	# Paste, [tab], paste, [tab][tab], paste, [tab], paste, [tab] [tab] [tab], paste, NEW LINE
	for sublist in ListContents:
		pyautogui.write(sublist[0])
		# pyautogui.typewrite(['tab'], interval=1)
		pyautogui.press('tab')
		pyautogui.write(sublist[1])
		pyautogui.press(['tab','tab'], interval=0.2)
		pyautogui.write(sublist[2])
		pyautogui.press('tab')
		pyautogui.write(sublist[3])
		pyautogui.press(['tab','tab','tab'], interval=0.2)
		pyautogui.write('US')
		pyautogui.press('tab') # 'tab' or 'enter'

fname = find_fname()
SheetContents = make_list_from_sheet(fname)

SheetContents = remove_nones(SheetContents)
SheetContents = convert_to_strings(SheetContents, 3)
SheetContents = formula_to_int(SheetContents)
SheetContents = cleanup_list(SheetContents)


copy_data_to_sap(SheetContents)

total_items  = len(SheetContents)
FILENAMES    = [SCRIPT_NAME, fname]
# Run Honeywell logger to write logs to file for YOY cost savings.
honlogger.write_log(DEPT, FILENAMES, LOG_START, NOUN_VERB, total_items, MIN_PER_ITEM)	

pyautogui.keyDown('alt')
pyautogui.typewrite(['tab'], interval=1)	# If IDE is not open, 
pyautogui.keyUp('alt')							# this should be edited.


print('\033[0;37;42m Script is now complete.')
